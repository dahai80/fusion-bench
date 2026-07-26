"""Report generator — produces JSON, Markdown, PDF, Excel, and chart reports from benchmark results.

Importers/callers: cli.py export commands, api/app.py /results/{id}/export endpoint.
Affected API: adds pdf/xlsx format options to export endpoint.
Data schema: reuses BenchmarkResult; PDF uses reportlab, Excel uses openpyxl.
User instruction: "对比PRD、架构、计划文档，查看是否还存在遗留、defer的任务" (P2-05/P2-06 PDF+Excel export).
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from ..engine.benchmark import BenchmarkResult

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates formatted reports from benchmark results."""

    @staticmethod
    def to_json(results: list[BenchmarkResult], filepath: str = "") -> str:
        """Export benchmark results as JSON."""
        data = {
            "generated_at": datetime.now().isoformat(),
            "total_benchmarks": len(results),
            "results": [r.metrics.to_dict() if hasattr(r, "metrics") else r for r in results],
        }
        output = json.dumps(data, indent=2, ensure_ascii=False)
        if filepath:
            Path(filepath).write_text(output, encoding="utf-8")
        return output

    @staticmethod
    def to_markdown(results: list[BenchmarkResult], title: str = "Benchmark Report") -> str:
        """Generate a Markdown report from benchmark results."""
        lines = [
            f"# {title}",
            "",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## Summary",
            "",
            "| Model | Config | Decode Speed | Prefill Speed | Peak Memory | Stable |",
            "|-------|--------|-------------|--------------|-------------|--------|",
        ]

        for r in results:
            model = r.model
            cfg = json.dumps(r.config, ensure_ascii=False) if r.config else "default"
            speed = f"{r.metrics.decode_speed:.1f} tok/s" if r.metrics.decode_speed else "N/A"
            prefill = f"{r.metrics.prefill_speed:.1f} tok/s" if r.metrics.prefill_speed else "N/A"
            mem = f"{r.metrics.peak_memory_mb:.0f} MB" if r.metrics.peak_memory_mb else "N/A"
            stable = "✅" if r.stable else "❌"
            lines.append(f"| {model} | {cfg} | {speed} | {prefill} | {mem} | {stable} |")

        lines.extend(
            [
                "",
                "## Speed Rankings",
                "",
            ]
        )

        sorted_results = sorted(results, key=lambda r: r.metrics.decode_speed, reverse=True)
        for i, r in enumerate(sorted_results, 1):
            if r.metrics.decode_speed:
                lines.append(f"{i}. **{r.model}**: {r.metrics.decode_speed:.1f} tok/s")

        lines.extend(
            [
                "",
                "## Configuration Details",
                "",
            ]
        )
        for r in results:
            lines.append(f"### {r.model}")
            lines.append("```json")
            lines.append(json.dumps(r.metrics.to_dict(), indent=2, ensure_ascii=False))
            lines.append("```")
            lines.append("")

        return "\n".join(lines)

    @staticmethod
    def generate_chart_path(results: list[BenchmarkResult], output_path: str = "") -> str:
        """Generate a speed comparison chart (PNG) using matplotlib."""
        if not results:
            return ""
        try:
            import matplotlib

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import numpy as np

            models = [r.model for r in results]
            speeds = [r.metrics.decode_speed for r in results]
            mems = [r.metrics.peak_memory_mb for r in results]

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

            # Speed chart
            colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(models)))
            bars1 = ax1.bar(range(len(models)), speeds, color=colors)
            ax1.set_xlabel("Model")
            ax1.set_ylabel("Decode Speed (tok/s)")
            ax1.set_title("Speed Comparison")
            ax1.set_xticks(range(len(models)))
            ax1.set_xticklabels(models, rotation=45, ha="right", fontsize=9)
            for bar, speed in zip(bars1, speeds, strict=False):
                ax1.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    f"{speed:.1f}",
                    ha="center",
                    va="bottom",
                    fontsize=8,
                )

            # Memory chart
            bars2 = ax2.bar(range(len(models)), mems, color=colors)
            ax2.set_xlabel("Model")
            ax2.set_ylabel("Peak Memory (MB)")
            ax2.set_title("Memory Usage Comparison")
            ax2.set_xticks(range(len(models)))
            ax2.set_xticklabels(models, rotation=45, ha="right", fontsize=9)
            for bar, mem in zip(bars2, mems, strict=False):
                ax2.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    f"{mem:.0f}",
                    ha="center",
                    va="bottom",
                    fontsize=8,
                )

            plt.tight_layout()
            path = output_path or "benchmark_chart.png"
            plt.savefig(path, dpi=150, bbox_inches="tight")
            plt.close()
            return path
        except ImportError:
            return ""

    @staticmethod
    def generate_config_template(result: BenchmarkResult) -> str:
        """Generate a fusion-mlx config template from the best result."""
        cfg = result.config
        return json.dumps(
            {
                "model": result.model,
                "inference": {
                    "max_tokens": cfg.get("max_tokens", 4096),
                    "temperature": cfg.get("temperature", 0.7),
                    "batch_size": cfg.get("batch_size", 1),
                },
                "performance": {
                    "expected_decode_speed": f"{result.metrics.decode_speed:.1f} tok/s",
                    "expected_peak_memory": f"{result.metrics.peak_memory_mb:.0f} MB",
                    "max_stable_context": result.max_stable_context,
                },
                "generated_by": "fusion-bench",
            },
            indent=2,
            ensure_ascii=False,
        )

    @staticmethod
    def to_pdf(
        results: list[BenchmarkResult],
        filepath: str = "",
        title: str = "Benchmark Report",
    ) -> str:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            logger.warning("reportlab not installed; pip install fusion-bench[pdf]")
            return ""

        path = filepath or "benchmark_report.pdf"
        doc = SimpleDocTemplate(path, pagesize=A4)
        styles = getSampleStyleSheet()
        elems = []
        elems.append(Paragraph(title, styles["Title"]))
        elems.append(
            Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles["Normal"],
            )
        )
        elems.append(Spacer(1, 12))

        table_data = [
            [
                "Model",
                "Config",
                "Decode (tok/s)",
                "Prefill (tok/s)",
                "Memory (MB)",
                "Stable",
            ]
        ]
        for r in results:
            cfg = json.dumps(r.config, ensure_ascii=False) if r.config else "default"
            ds = f"{r.metrics.decode_speed:.1f}" if r.metrics.decode_speed else "N/A"
            ps = f"{r.metrics.prefill_speed:.1f}" if r.metrics.prefill_speed else "N/A"
            mm = f"{r.metrics.peak_memory_mb:.0f}" if r.metrics.peak_memory_mb else "N/A"
            st = "Yes" if r.stable else "No"
            table_data.append([r.model, cfg, ds, ps, mm, st])

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elems.append(table)

        sorted_results = sorted(results, key=lambda r: r.metrics.decode_speed, reverse=True)
        elems.append(Spacer(1, 12))
        elems.append(Paragraph("Speed Rankings", styles["Heading2"]))
        for i, r in enumerate(sorted_results, 1):
            if r.metrics.decode_speed:
                elems.append(
                    Paragraph(
                        f"{i}. {r.model}: {r.metrics.decode_speed:.1f} tok/s",
                        styles["Normal"],
                    )
                )

        doc.build(elems)
        logger.info("PDF report saved to %s", path)
        return path

    @staticmethod
    def to_excel(
        results: list[BenchmarkResult],
        filepath: str = "",
        sheet_name: str = "Benchmarks",
    ) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed; pip install fusion-bench[excel]")
            return ""

        path = filepath or "benchmark_report.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = [
            "Model",
            "Config",
            "Decode Speed (tok/s)",
            "Prefill Speed (tok/s)",
            "Peak Memory (MB)",
            "Prompt Tokens",
            "Completion Tokens",
            "Stable",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in results:
            cfg = json.dumps(r.config, ensure_ascii=False) if r.config else "default"
            ws.append(
                [
                    r.model,
                    cfg,
                    r.metrics.decode_speed or 0,
                    r.metrics.prefill_speed or 0,
                    r.metrics.peak_memory_mb or 0,
                    r.metrics.prompt_tokens or 0,
                    r.metrics.completion_tokens or 0,
                    "Yes" if r.stable else "No",
                ]
            )

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        wb.save(path)
        logger.info("Excel report saved to %s", path)
        return path

    @staticmethod
    def to_html(
        results: list[BenchmarkResult],
        filepath: str = "",
        title: str = "Benchmark Report",
    ) -> str:
        generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = ""
        for idx, r in enumerate(results, 1):
            cfg = json.dumps(r.config, ensure_ascii=False) if r.config else "default"
            ds = f"{r.metrics.decode_speed:.1f}" if r.metrics.decode_speed else "N/A"
            ps = f"{r.metrics.prefill_speed:.1f}" if r.metrics.prefill_speed else "N/A"
            mm = f"{r.metrics.peak_memory_mb:.0f}" if r.metrics.peak_memory_mb else "N/A"
            badge = '<span style="color:green">Yes</span>' if r.stable else '<span style="color:red">No</span>'
            rows += f"<tr><td>{idx}</td><td>{r.model}</td><td>{cfg}</td><td>{ds}</td><td>{ps}</td><td>{mm}</td><td>{badge}</td></tr>\n"
        rankings = ""
        for _i, r in enumerate(sorted(results, key=lambda x: x.metrics.decode_speed, reverse=True), 1):
            if r.metrics.decode_speed:
                rankings += f"<li><strong>{r.model}</strong>: {r.metrics.decode_speed:.1f} tok/s</li>\n"
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; margin: 2rem auto; max-width: 960px; color: #333; }}
h1 {{ border-bottom: 2px solid #4472C4; padding-bottom: 0.5rem; }}
table {{ border-collapse: collapse; width: 100%; margin: 1rem 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: left; }}
th {{ background: #4472C4; color: #fff; }}
tr:nth-child(even) {{ background: #f9f9f9; }}
footer {{ margin-top: 2rem; font-size: 0.85rem; color: #888; }}
</style>
</head>
<body>
<h1>{title}</h1>
<p>Generated: {generated}</p>
<h2>Summary</h2>
<table>
<thead><tr><th>#</th><th>Model</th><th>Config</th><th>Decode (tok/s)</th><th>Prefill (tok/s)</th><th>Memory (MB)</th><th>Stable</th></tr></thead>
<tbody>{rows}</tbody>
</table>
<h2>Speed Rankings</h2>
<ol>{rankings}</ol>
<footer>Generated by Fusion-Bench</footer>
</body>
</html>"""
        if filepath:
            Path(filepath).write_text(html, encoding="utf-8")
        logger.info("HTML report saved to %s", filepath or "<string>")
        return html

    @staticmethod
    def generate_radar_chart(
        results: list[BenchmarkResult],
        output_path: str = "",
        title: str = "Benchmark Radar",
    ) -> str:
        try:
            import matplotlib

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import numpy as np
        except ImportError:
            logger.warning("matplotlib not installed; pip install matplotlib")
            return ""

        if not results:
            return ""

        categories = [
            "Decode Speed",
            "Prefill Speed",
            "Memory Eff.",
            "Stability",
            "Context Len",
        ]
        angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
        angles += angles[:1]

        fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
        colors = plt.cm.Set2(np.linspace(0, 1, max(len(results), 1)))

        max_decode = (
            max(
                (r.metrics.decode_speed for r in results if r.metrics.decode_speed),
                default=1,
            )
            or 1
        )
        max_prefill = (
            max(
                (r.metrics.prefill_speed for r in results if r.metrics.prefill_speed),
                default=1,
            )
            or 1
        )
        max_mem = (
            max(
                (r.metrics.peak_memory_mb for r in results if r.metrics.peak_memory_mb),
                default=1,
            )
            or 1
        )
        max_ctx = (
            max(
                (r.max_stable_context for r in results if r.max_stable_context),
                default=1,
            )
            or 1
        )

        for idx, r in enumerate(results):
            decode_norm = (r.metrics.decode_speed or 0) / max_decode
            prefill_norm = (r.metrics.prefill_speed or 0) / max_prefill
            mem_norm = 1.0 - min((r.metrics.peak_memory_mb or 0) / max_mem, 1.0)
            stability = 1.0 if r.stable else 0.3
            ctx_norm = min((r.max_stable_context or 0) / max_ctx, 1.0) if max_ctx > 0 else 0.0
            values = [decode_norm, prefill_norm, mem_norm, stability, ctx_norm]
            values += values[:1]
            ax.plot(angles, values, "o-", linewidth=1.5, label=r.model, color=colors[idx])
            ax.fill(angles, values, alpha=0.15, color=colors[idx])

        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories, fontsize=10)
        ax.set_ylim(0, 1.1)
        ax.set_title(title, pad=20, fontsize=13)
        ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1), fontsize=9)

        path = output_path or "benchmark_radar.png"
        plt.savefig(path, dpi=150, bbox_inches="tight")
        plt.close()
        logger.info("Radar chart saved to %s", path)
        return path

    @staticmethod
    def generate_trend_chart(
        trend_data: list[dict[str, Any]],
        output_path: str = "",
        title: str = "Quality Trend",
    ) -> str:
        try:
            import matplotlib

            matplotlib.use("Agg")
            from datetime import datetime as dt

            import matplotlib.pyplot as plt
            import numpy as np
            from matplotlib.dates import DateFormatter
        except ImportError:
            logger.warning("matplotlib not installed; pip install matplotlib")
            return ""

        if not trend_data:
            return ""

        by_model: dict[str, list[dict]] = {}
        for point in trend_data:
            model = point.get("model", "unknown")
            by_model.setdefault(model, []).append(point)

        fig, ax = plt.subplots(figsize=(12, 5))
        colors = plt.cm.Set2(np.linspace(0, 1, max(len(by_model), 1)))

        for idx, (model, points) in enumerate(by_model.items()):
            points.sort(key=lambda p: p.get("timestamp", ""))
            timestamps = [dt.fromisoformat(p["timestamp"][:19]) for p in points if "timestamp" in p]
            values = [p.get("metric_value", 0) for p in points]
            if timestamps:
                ax.plot(
                    timestamps,
                    values,
                    "o-",
                    label=model,
                    color=colors[idx],
                    markersize=4,
                )

        ax.set_xlabel("Time")
        ax.set_ylabel("Metric Value")
        ax.set_title(title)
        ax.legend(fontsize=9)
        ax.xaxis.set_major_formatter(DateFormatter("%m-%d %H:%M"))
        fig.autofmt_xdate()
        plt.tight_layout()

        path = output_path or "benchmark_trend.png"
        plt.savefig(path, dpi=150, bbox_inches="tight")
        plt.close()
        logger.info("Trend chart saved to %s", path)
        return path
