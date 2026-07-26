"""Report export enhancements - PDF, Excel, radar chart, trend chart.

Importers/callers: reporter/report.py calls to_pdf()/to_excel()/radar_chart()/trend_chart().
Affected API: extends ReportGenerator with PDF/Excel export; no new REST endpoints.
Data schema: reuses existing EvalResult.to_dict()/SuiteResult dict; no schema changes.
User instruction: "把所有没完全做完，没启动做，有差距的全部启动补齐" (P1-4 report export + Enhancement C visualization).
"""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def to_pdf(data: dict[str, Any], output_path: str | Path) -> str:
    """Export benchmark results to PDF using reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        logger.error("reportlab not installed. Run: pip install reportlab")
        return _fallback_text(data, output_path)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(str(output_path), pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Fusion-Bench Report", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Suite: {data.get('suite_id', 'N/A')}", styles["Normal"]))
    elements.append(Paragraph(f"Model: {data.get('model', 'N/A')}", styles["Normal"]))
    elements.append(Paragraph(f"Level: {data.get('level', 'N/A')}", styles["Normal"]))
    elements.append(Paragraph(f"Passed: {'Yes' if data.get('overall_passed') else 'No'}", styles["Normal"]))
    elements.append(Paragraph(f"Duration: {data.get('duration_seconds', 0):.1f}s", styles["Normal"]))
    elements.append(Spacer(1, 12))

    results = data.get("results", [])
    if results:
        table_data = [["Task", "Executor", "Metric", "Value", "Pass Rate"]]
        for r in results:
            if isinstance(r, dict):
                table_data.append(
                    [
                        str(r.get("task_id", ""))[:20],
                        str(r.get("executor_key", ""))[:12],
                        str(r.get("metric_name", ""))[:15],
                        f"{r.get('metric_value', 0):.4f}",
                        f"{r.get('pass_rate', 0):.2%}",
                    ]
                )
        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    logger.info("PDF exported: %s", output_path)
    return str(output_path)


def _fallback_text(data: dict[str, Any], output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.warning("reportlab unavailable, wrote JSON instead: %s", output_path)
    return str(output_path)


def to_excel(data: dict[str, Any], output_path: str | Path) -> str:
    """Export benchmark results to Excel using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return _fallback_csv(data, output_path)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(["Suite ID", "Model", "Level", "Overall Passed", "Duration (s)"])
    ws.append(
        [
            data.get("suite_id", ""),
            data.get("model", ""),
            data.get("level", ""),
            "Yes" if data.get("overall_passed") else "No",
            f"{data.get('duration_seconds', 0):.1f}",
        ]
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    results = data.get("results", [])
    if results:
        ws2 = wb.create_sheet("Results")
        ws2.append(["Task ID", "Executor", "Metric", "Value", "Pass Rate", "Errors"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        for r in results:
            if isinstance(r, dict):
                ws2.append(
                    [
                        r.get("task_id", ""),
                        r.get("executor_key", ""),
                        r.get("metric_name", ""),
                        r.get("metric_value", 0),
                        r.get("pass_rate", 0),
                        "; ".join(r.get("errors", []))[:100],
                    ]
                )

    wb.save(str(output_path))
    logger.info("Excel exported: %s", output_path)
    return str(output_path)


def _fallback_csv(data: dict[str, Any], output_path: str | Path) -> str:
    output_path = Path(output_path).with_suffix(".csv")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["key", "value"])
        for k, v in data.items():
            if not isinstance(v, (list, dict)):
                writer.writerow([k, v])
    logger.warning("openpyxl unavailable, wrote CSV instead: %s", output_path)
    return str(output_path)


def radar_chart(metrics: dict[str, float], output_path: str | Path, title: str = "Model Performance") -> str:
    """Generate a radar/spider chart from metric key-value pairs."""
    try:
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError:
        logger.error("matplotlib/numpy not installed")
        return ""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    labels = list(metrics.keys())
    values = list(metrics.values())
    n = len(labels)
    if n < 3:
        logger.warning("Radar chart needs >= 3 metrics, got %d", n)
        return ""

    angles = np.linspace(0, 2 * np.pi, n, endpoint=False).tolist()
    values_plot = values + [values[0]]
    angles += [angles[0]]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
    ax.plot(angles, values_plot, "o-", linewidth=2)
    ax.fill(angles, values_plot, alpha=0.25)
    ax.set_thetagrids(np.degrees(angles[:-1]), labels)
    ax.set_title(title, size=14, pad=20)
    ax.set_ylim(0, 1)

    fig.savefig(str(output_path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    logger.info("Radar chart saved: %s", output_path)
    return str(output_path)


def trend_chart(
    series: list[dict[str, Any]],
    metric_key: str,
    output_path: str | Path,
    title: str = "Metric Trend",
) -> str:
    """Generate a trend line chart from a series of data points."""
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        logger.error("matplotlib not installed")
        return ""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    x_labels = []
    y_values = []
    for point in series:
        x_labels.append(point.get("label", point.get("version", str(len(x_labels)))))
        y_values.append(point.get(metric_key, 0))

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(x_labels, y_values, "o-", linewidth=2, markersize=6)
    ax.set_xlabel("Version / Run")
    ax.set_ylabel(metric_key)
    ax.set_title(title)
    ax.grid(True, alpha=0.3)
    plt.xticks(rotation=45, ha="right")

    fig.savefig(str(output_path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    logger.info("Trend chart saved: %s", output_path)
    return str(output_path)
